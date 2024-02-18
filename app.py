from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook, load_workbook
from pathlib import Path 
from urllib.parse import quote
from time import sleep
from datetime import datetime, timedelta
import os
import sys
import locale



def planinha_atualizada():
    arquivo = 'Planilha Atualizada.xlsx'

    if os.path.exists(arquivo):
        menu()

    os.system('cls')

    pasta = Path('Planilha MK-AUTH')
    
    if not pasta.exists():
        pasta.mkdir()
        print('Pasta não encontrada')
        sleep(0.5)

        print('Criando nova pasta...')
        sleep(0.5)

        print('Pasta criada com sucesso!')
        sleep(0.5)
        pass


    diretorio = "Planilha MK-AUTH/"

    # Listar todos os arquivos no diretório
    arquivos = os.listdir(diretorio)

    arquivo_xlsx = None
    
    for arquivo in arquivos:
        if arquivo.endswith(".xlsx"):
            arquivo_xlsx = os.path.join(diretorio, arquivo)
            break


    if arquivo_xlsx:
        try:
            print("Planilha encontrada!")
            sleep(1)
            
            print('Criando uma nova planilha atualizada...')
            sleep(1)

            workbook = openpyxl.load_workbook(f'{arquivo_xlsx}')
    
            pagina_clientes = workbook['Planilha1']

            wb = Workbook()
            ws = wb.active

            ws.append(['Nome', 'Número', 'Vencimento'])


            for id, linha in enumerate(pagina_clientes.iter_rows(min_row=3)):
                
                print(f'{id}: {linha[1].value}| Número: {linha[15].value} | Vencimento: {linha[25].value}')


                nome = linha[1].value
                numero = linha[15].value
                vencimento = linha[25].value


                ws.append([f'{nome}', f'{numero}', f'{vencimento}'])

            sleep(2)
            
            os.system('cls')
            print('Planilha atualizada criada com sucesso!')
            wb.save('Planilha Atualizada.xlsx')
            sleep(3)


            # Carregando a planilha original e a planilha cópia
            wb_original = load_workbook(f'{arquivo_xlsx}')
            
            wb_copia = load_workbook('Planilha Atualizada.xlsx')

            # Selecionando as folhas ativas
            ws_original = wb_original.active
            ws_copia = wb_copia.active

            ws_copia.column_dimensions['A'].width = 40
            ws_copia.column_dimensions['B'].width = 20
            ws_copia.column_dimensions['C'].width = 15

            wb_copia.save('Planilha Atualizada.xlsx')

        # except IndexError as error:
        #     os.system('cls')
        #     print('Erro:')
        #     print(f'A planilha {arquivo_xlsx} não tem uma ou mais colunas com os dados necessários.')
        #     print('\nSão elas: Coluna 1 | Coluna 15 | Coluna 25')
        #     print('Verifique se as mesmas existem')
        #     print('\nObs: a planilha do sistema MK-AUTH vem com as colunas corretas, veja se a mesma está na pasta.')
        #     input('\nPresioner ENTER para fechar.')
        #     sys.exit()


        except KeyError:
            os.system('cls')
            print('Erro:')
            print('O WorkSheet dentro da planilha precisa ser renomeada para Planilha1 ')
            input('Presioner ENTER para fechar.')
            sys.exit()

    else:
        os.system('cls')
        print('Nenhum Arquivo.xlsx Foi Encontrado na Pasta Planilha MK-AUTH')
        sleep(1)

        print('\nAdicione a planilha do MK-AUTH na pasta Planilha MK-AUTH e tente novamente.')
        input('\nPressione ENTER para fechar.')
        sys.exit()


def menu():
    
    pasta = Path('Não Enviados')
    if not pasta.exists():
        pasta.mkdir()

    planilha = Path('Não Enviados/Planilha de Reenvio.xlsx')

    if not planilha.exists():
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

        wb.save('Não Enviados/Planilha de Reenvio.xlsx')
    
    while True:
        os.system('cls')
        print('''WhatsApp Bot de Mensagem Automática

(1) Ativar Mensagem Automática
(2) Tentar Reenviar Mensagens 
(3) Programar para domingo
              
                    ''')

        opcao = input(f'Digite 1 Para Ativar o Bot: ')

        if opcao == '1':
            os.system('cls')
            print('PARA CANCELAR FECHE O APP...')
            sleep(3)            
            mensagem_automatica()
            break
        
        if opcao == '2':
            os.system('cls')
            print('PARA CANCELAR FECHE O APP...')
            sleep(3)  
            reenviar_mensagem()

        if opcao == '3':
            domingo()
        
        else:
            continue


def mensagem_automatica():

    current_directory = os.path.dirname(os.path.abspath(__file__))

    session_data_directory = os.path.join(current_directory, "session_data")

    # Verifica se a pasta de sessão existe, senão a cria
    if not os.path.exists(session_data_directory):
        os.makedirs(session_data_directory)

    chrome_options = Options()

    chrome_options.add_argument(f'--user-data-dir={session_data_directory}')
    chrome_options.add_argument('--profile-directory=Default')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-popup-blocking')

    chrome_options.add_experimental_option("detach", True)
    servico = Service(ChromeDriverManager().install())




    workbook = openpyxl.load_workbook('Planilha Atualizada.xlsx')
    pagina_clientes = workbook['Sheet']

    # Extrai todos os dados da planilha cópia para o envio das mensagens
    for linha in pagina_clientes.iter_rows(min_row=2):
        navegador = webdriver.Chrome(service=servico, options=chrome_options)

        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value

        if vencimento is not None:
            
            data_antecipada = timedelta(days=int(vencimento)) - timedelta(days=1)        
            data_atual = datetime.now().day

            # Faz a verificação da data, caso seja um dia antes do vencimento ele enviará a mensagem, senão irá ignorar
            if data_antecipada.days == data_atual:

                # Caso o número seja vazio ele é alertado e registrado na planilha "Planilha de Reenvio"
                if telefone is None or telefone == '':
                    
                    print(f'Não foi possível enviar a mensagem para {nome} | (Sem Número)')

                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    linha = 1
                    coluna = 1
                    
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    # if sheet.cell(row=linha, column=coluna).value is None:
                    dados = [f'{nome}', 'Sem Número', f'{vencimento}']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
                    continue

                # Mensagem de exemplo que os clientes receberá contendo o nome e o vencimento. Sendo possível a troca a mensagem para a que mais agradar
                mensagem = f'''*Mensagem Automática:*

Olá {nome.title()} seu boleto vence dia {vencimento} (amanhã). Venha pagar presencialmente ou utilize nossos meios de pagamento:

Pix CNPJ: 26.752.862/0001-64 | Plnalto Telecom

Conta para depósito: 

Caixa Econômica Federal : 3880 1288 000981858801-6 Marlene de Jesus Coelho

*Não se esqueça de nos enviar o comprovante!*

Caso o pagamento já tenha sido efetuado, desconsidere esta mensagem.'''
                try:
                    navegador.get(f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}')

                    #Da ate 30 segundos para o botão ficar disponível
                    botao_enviar = WebDriverWait(navegador, 60).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span'))
                    )

                    botao_enviar.click()

                    # sleep(1.5)
                    # navegador.close()


                except Exception as error:
                    print(f'Não foi possível enviar a mensagem para {nome} | (Número Inválido) {error}')
                    
                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    linha = 1
                    coluna = 1
                    
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    dados = [f'{nome}', 'Número Inválido', f'{vencimento}']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')

                finally:
                    sleep(1.5)
                    navegador.quit()
    
    locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')
    data_atual = datetime.now()
    dia_por_extenso = data_atual.strftime('%A')
    
    if dia_por_extenso == 'domingo':
        sleep(2)
        os.system("shutdown /s /t 1")

    print('\nMensagens enviadas com sucesso!')
    input('Pressione ENTER para voltar')
    menu()


def reenviar_mensagem():
    
    current_directory = os.path.dirname(os.path.abspath(__file__))

    session_data_directory = os.path.join(current_directory, "session_data")

    # Verifica se a pasta de sessão existe, senão a cria
    if not os.path.exists(session_data_directory):
        os.makedirs(session_data_directory)

    chrome_options = Options()

    chrome_options.add_argument(f'--user-data-dir={session_data_directory}')
    chrome_options.add_argument('--profile-directory=Default')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-popup-blocking')

    chrome_options.add_experimental_option("detach", True)
    servico = Service(ChromeDriverManager().install())

    navegador = webdriver.Chrome(service=servico, options=chrome_options)



    workbook = openpyxl.load_workbook('Planilha Atualizada.xlsx')
    pagina_clientes = workbook['Sheet']

    # Extrai todos os dados da planilha cópia para o envio das mensagens
    for index, linha in enumerate(pagina_clientes.iter_rows(min_row=2)):

        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value

        if vencimento is not None:
            
            data_antecipada = timedelta(days=int(vencimento)) - timedelta(days=1)        
            data_atual = datetime.now().day

                    # Faz a verificação da data, caso seja um dia antes do vencimento ele enviará a mensagem, senão irá ignorar
            if data_antecipada.days == data_atual:

                # Caso o número seja vazio ele é alertado e registrado na planilha "Planilha de Reenvio"
                if telefone is None or telefone == '':
                    
                    print(f'Não foi possível enviar a mensagem para {nome} | (Sem Número)')

                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    sheet.delete_rows(index)

                    linha = 1
                    coluna = 1
                    
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    # if sheet.cell(row=linha, column=coluna).value is None:
                    dados = [f'{nome}', 'Sem Número', f'{vencimento}']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
                    continue

                # Mensagem de exemplo que os clientes receberá contendo o nome e o vencimento. Sendo possível a troca a mensagem para a que mais agradar
                mensagem = f'''*Mensagem Automática:*

Olá {nome.title()} seu boleto vence dia {vencimento} (amanhã). Venha pagar presencialmente ou utilize nossos meios de pagamento:

Pix CNPJ: 26.752.862/0001-64 | Plnalto Telecom

Conta para depósito: 

Caixa Econômica Federal : 3880 1288 000981858801-6 Marlene de Jesus Coelho

*Não se esqueça de nos enviar o comprovante!*

Caso o pagamento já tenha sido efetuado, desconsidere esta mensagem.'''
                try:
                    navegador.get(f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}')

                    #Da ate 30 segundos para o botão ficar disponível
                    botao_enviar = WebDriverWait(navegador, 60).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span'))
                    )

                    botao_enviar.click()

                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    
                    sheet = workbook.active
                    sheet.delete_rows(index)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')

                    sleep(1.5)
                    navegador.close()


                except Exception as error:
                    os.system('cls')
                    print(f'Não foi possível enviar a mensagem para {nome} | (Número Inválido)')
                    
                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    linha = 1
                    coluna = 1
                    
                    sheet.delete_rows(index)

                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    dados = [f'{nome}', 'Número Inválido', f'{vencimento}']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')

        input('Pressione ENTER para voltar')
        main()


def domingo():
    os.system('cls')
    dia = ''

    print('Programando Bot...')
    sleep(2)

    print('Bot programado com sucesso!')
    sleep(2)

    print('Iniciando Bot...')
    sleep(2)

    os.system('cls')
    while dia != 'domingo':    
        locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')
        
        data_atual = datetime.now()
        
        horario_atual = datetime.now().strftime('%H:%M:%S')
        
        horario_especifico = datetime.strptime("07:00:00", "%H:%M:%S").strftime('%H:%M:%S')
        
        dia_por_extenso = data_atual.strftime('%A')
        

        if dia_por_extenso == 'domingo' and horario_atual >= horario_especifico:
            os.system('cls')
            print('Iniciando...')
            sleep(3)
            dia = 'domingo'
            mensagem_automatica()
        
        print('Assim que for domingo o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
        print('Não feche o programa!')
        print('\nEsperando.')
        sleep(1)
        os.system('cls')

        print('Assim que for domingo o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
        print('Não feche o programa!')
        print('\nEsperando..')
        sleep(1)
        os.system('cls')

        print('Assim que for domingo o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
        print('Não feche o programa!')
        print('\nEsperando...')
        sleep(1)
        os.system('cls')    


def main():
    planinha_atualizada()
    menu()

if __name__=='__main__':
    main()